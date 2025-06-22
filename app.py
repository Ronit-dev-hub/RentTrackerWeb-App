import os
import logging
from flask import Flask, render_template, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.orm import DeclarativeBase
from werkzeug.middleware.proxy_fix import ProxyFix
from datetime import datetime
from decimal import Decimal

# Configure logging
logging.basicConfig(level=logging.DEBUG)

class Base(DeclarativeBase):
    pass

db = SQLAlchemy(model_class=Base)

# Create the app
app = Flask(__name__)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)

# Configure the database
database_url = os.environ.get("DATABASE_URL")
if not database_url:
    raise RuntimeError("DATABASE_URL environment variable is not set")

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_recycle": 300,
    "pool_pre_ping": True,
}

# Initialize the app with the extension
db.init_app(app)

with app.app_context():
    # Import models here so tables are created
    import models
    db.create_all()

@app.route('/')
def index():
    """Main application page"""
    return render_template('index.html')

@app.route('/api/properties', methods=['GET'])
def get_properties():
    """Get all properties"""
    try:
        properties = models.Property.query.all()
        return jsonify([prop.to_dict() for prop in properties])
    except Exception as e:
        logging.error(f"Error fetching properties: {e}")
        return jsonify({'error': 'Failed to fetch properties'}), 500

@app.route('/api/properties', methods=['POST'])
def create_property():
    """Create a new property"""
    try:
        data = request.get_json()
        
        # Parse lease start date
        lease_start_date = datetime.strptime(data['leaseStartDate'], '%Y-%m-%d').date()
        initial_rent = Decimal(str(data['initialRent']))
        
        property = models.Property(
            name=data['name'],
            address=data['address'],
            renter_name=data['renterName'],
            renter_contact=data.get('renterContact', ''),
            initial_rent=initial_rent,
            current_rent=initial_rent,  # Initially same as initial rent
            lease_start_date=lease_start_date,
            expected_rent_date=data['expectedRentDate'],
            yearly_increase_percent=Decimal(str(data['yearlyIncreasePercent'])) if data.get('yearlyIncreasePercent') else None,
            yearly_increase_amount=Decimal(str(data['yearlyIncreaseAmount'])) if data.get('yearlyIncreaseAmount') else None
        )
        
        db.session.add(property)
        db.session.commit()
        
        # Calculate current rent based on lease duration
        property.calculate_current_rent()
        
        return jsonify(property.to_dict()), 201
    except Exception as e:
        logging.error(f"Error creating property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create property'}), 500

@app.route('/api/properties/<int:property_id>', methods=['PUT'])
def update_property(property_id):
    """Update a property"""
    try:
        property = models.Property.query.get_or_404(property_id)
        data = request.get_json()
        
        property.name = data['name']
        property.address = data['address']
        property.renter_name = data['renterName']
        property.renter_contact = data.get('renterContact', '')
        property.initial_rent = Decimal(str(data['initialRent']))
        property.lease_start_date = datetime.strptime(data['leaseStartDate'], '%Y-%m-%d').date()
        property.expected_rent_date = data['expectedRentDate']
        property.yearly_increase_percent = Decimal(str(data['yearlyIncreasePercent'])) if data.get('yearlyIncreasePercent') else None
        property.yearly_increase_amount = Decimal(str(data['yearlyIncreaseAmount'])) if data.get('yearlyIncreaseAmount') else None
        
        # Recalculate current rent
        property.calculate_current_rent()
        
        db.session.commit()
        
        return jsonify(property.to_dict())
    except Exception as e:
        logging.error(f"Error updating property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update property'}), 500

@app.route('/api/properties/<int:property_id>', methods=['DELETE'])
def delete_property(property_id):
    """Delete a property"""
    try:
        property = models.Property.query.get_or_404(property_id)
        db.session.delete(property)
        db.session.commit()
        
        return jsonify({'message': 'Property deleted successfully'})
    except Exception as e:
        logging.error(f"Error deleting property: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete property'}), 500

@app.route('/api/rent-records', methods=['GET'])
def get_rent_records():
    """Get all rent records"""
    try:
        records = models.RentRecord.query.all()
        result = {}
        
        for record in records:
            key = f"{record.year}-{record.month}"
            if key not in result:
                result[key] = {}
            result[key][str(record.property_id)] = record.to_dict()
        
        return jsonify(result)
    except Exception as e:
        logging.error(f"Error fetching rent records: {e}")
        return jsonify({'error': 'Failed to fetch rent records'}), 500

@app.route('/api/rent-records', methods=['POST'])
def update_rent_record():
    """Update or create a rent record"""
    try:
        data = request.get_json()
        property_id = data['propertyId']
        year = data['year']
        month = data['month']
        
        # Try to find existing record
        record = models.RentRecord.query.filter_by(
            property_id=property_id,
            year=year,
            month=month
        ).first()
        
        if not record:
            record = models.RentRecord(
                property_id=property_id,
                year=year,
                month=month
            )
            db.session.add(record)
        
        # Update fields
        record.received = data.get('received', False)
        record.payment_mode = data.get('paymentMode', '')
        record.rent_amount = Decimal(str(data['rentAmount'])) if data.get('rentAmount') else None
        
        if data.get('expectedDate'):
            record.expected_date = datetime.strptime(data['expectedDate'], '%Y-%m-%d').date()
        else:
            record.expected_date = None
            
        if data.get('receivedDate'):
            record.received_date = datetime.strptime(data['receivedDate'], '%Y-%m-%d').date()
        else:
            record.received_date = None
        
        record.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify(record.to_dict())
    except Exception as e:
        logging.error(f"Error updating rent record: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update rent record'}), 500

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Export all data to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        import calendar
        
        wb = Workbook()
        
        # Properties sheet
        properties_sheet = wb.active
        properties_sheet.title = "Properties"
        
        # Headers for properties
        properties_headers = [
            'Property ID', 'Property Name', 'Address', 'Renter Name', 'Renter Contact',
            'Initial Rent', 'Current Rent', 'Lease Start Date', 'Expected Rent Day',
            'Yearly Increase %', 'Yearly Increase Amount', 'Created Date'
        ]
        
        for col, header in enumerate(properties_headers, 1):
            cell = properties_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Add property data
        properties = models.Property.query.all()
        for row, prop in enumerate(properties, 2):
            properties_sheet.cell(row=row, column=1, value=prop.id)
            properties_sheet.cell(row=row, column=2, value=prop.name)
            properties_sheet.cell(row=row, column=3, value=prop.address)
            properties_sheet.cell(row=row, column=4, value=prop.renter_name)
            properties_sheet.cell(row=row, column=5, value=prop.renter_contact or '')
            properties_sheet.cell(row=row, column=6, value=float(prop.initial_rent))
            properties_sheet.cell(row=row, column=7, value=float(prop.current_rent))
            properties_sheet.cell(row=row, column=8, value=prop.lease_start_date.isoformat())
            properties_sheet.cell(row=row, column=9, value=prop.expected_rent_date)
            properties_sheet.cell(row=row, column=10, value=float(prop.yearly_increase_percent) if prop.yearly_increase_percent else '')
            properties_sheet.cell(row=row, column=11, value=float(prop.yearly_increase_amount) if prop.yearly_increase_amount else '')
            properties_sheet.cell(row=row, column=12, value=prop.created_at.isoformat())
        
        # Rent Records sheet
        rent_sheet = wb.create_sheet("Rent Records")
        
        rent_headers = [
            'Property Name', 'Renter Name', 'Year', 'Month', 'Month Name',
            'Rent Amount', 'Expected Date', 'Received Date', 'Payment Mode',
            'Status', 'Days Late', 'Created Date'
        ]
        
        for col, header in enumerate(rent_headers, 1):
            cell = rent_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        
        # Add rent record data
        records = models.RentRecord.query.join(models.Property).all()
        for row, record in enumerate(records, 2):
            rent_sheet.cell(row=row, column=1, value=record.property.name)
            rent_sheet.cell(row=row, column=2, value=record.property.renter_name)
            rent_sheet.cell(row=row, column=3, value=record.year)
            rent_sheet.cell(row=row, column=4, value=record.month)
            rent_sheet.cell(row=row, column=5, value=calendar.month_name[record.month + 1])
            rent_sheet.cell(row=row, column=6, value=float(record.rent_amount) if record.rent_amount else float(record.property.current_rent))
            rent_sheet.cell(row=row, column=7, value=record.expected_date.isoformat() if record.expected_date else '')
            rent_sheet.cell(row=row, column=8, value=record.received_date.isoformat() if record.received_date else '')
            rent_sheet.cell(row=row, column=9, value=record.payment_mode or '')
            rent_sheet.cell(row=row, column=10, value='Received' if record.received else 'Pending')
            
            # Calculate days late
            days_late = ''
            if record.received and record.expected_date and record.received_date:
                days_diff = (record.received_date - record.expected_date).days
                if days_diff > 0:
                    days_late = days_diff
            
            rent_sheet.cell(row=row, column=11, value=days_late)
            rent_sheet.cell(row=row, column=12, value=record.created_at.isoformat())
        
        # Auto-adjust column widths
        for sheet in [properties_sheet, rent_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name=f'property_rent_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}")
        return jsonify({'error': 'Failed to export data'}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
